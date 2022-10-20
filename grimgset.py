import os, sys, os.path
from openpyxl import Workbook

from PyQt6.QtWidgets import *
from PyQt6.QtGui import QPixmap, QImage, QAction
from PyQt6.QtCore import (
    pyqtSlot, pyqtSignal
)

from ui.ui import Label_buttons, Settings_widget, Main_thread_ui


class Image_Procession(QMainWindow):
    ''' Main working thread. '''
    rename = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.name = None
        self.scaleFactor = 0.0

        # Dictionary for labels, paths, etc..
        self.labl_dic = {}
        self.output_img_set = os.getcwd() + '/output_img_set'  # Directory to save imgs

        Main_thread_ui.create_main_ui(self)
        self.statusBar()

        # white_black qss switcher
        self.white_black = True
        self.white_dark_clicked()

        self.counter_new_button_id = 0  # Counter for dynamicly added widgets

        self.count = 0  # Counter for files/images
        self.names = []  # List for names of images
        self.row_count = 0  # row_count for csv_processor

        # Counting buttons presses to rename images in corresponded folders
        self.count_button_1 = 0

        self.add_new_button()
        self.add_new_button()

        # Creating menus and populating them with actions
        self.createActions()
        self.createMenus()

        # Preset csv_table
        self.wb = Workbook()
        self.wb.save('img_data.xlsx')

        # Set up labels from start
        self.change_text()

    def about(self):
        ''' Invokes text box about Grimgset programm. '''
        QMessageBox.about(
            self, 'About Grimgset',
            '<p>The <b>Grimgset</b> is an app for manual labeling images </p>'
        )

    def createActions(self):
        ''' Create actions and corresponded triggers. '''
        self.open_act_multiple = QAction('&Open multiple...', self, triggered=self.open_files)
        self.exit_act = QAction('E&xit', self, shortcut='Ctrl+Q', triggered=self.close)
        self.about_act = QAction('&About', self, triggered=self.about)

    def createMenus(self):
        ''' Create menus and assing actions to them. '''
        self.file_menu = QMenu('&File', self)
        self.file_menu.addAction(self.open_act_multiple)
        self.file_menu.addSeparator()
        self.file_menu.addAction(self.exit_act)

        self.help_menu = QMenu('&Help', self)
        self.help_menu.addAction(self.about_act)

        self.menuBar().addMenu(self.file_menu)
        self.menuBar().addMenu(self.help_menu)

    def white_dark_clicked(self):
        ''' Changes white/black style. '''
        if self.white_black is True:
            with open('./styles/white_style.qss', 'r') as f:
                self.setStyleSheet(f.read())
                self.white_black = False
        else:
            with open('./styles/black_style.qss', 'r') as f:
                self.setStyleSheet(f.read())
                self.white_black = True

    def change_text(self):
        naming = 'cat'
        self.rename.emit(naming)

    @pyqtSlot(int, str)
    def create_dictionary(self, wid, text):
        '''
        Checks changes in set_wid labels.
        And creates labels dictionary.
        '''
        if wid not in self.labl_dic.keys():
            l_dic = {}
            upd = [['label', text]]
            l_dic.update(upd)
            upd_2 = [[wid, l_dic]]
            self.labl_dic.update(upd_2)
        else:
            l_dic = {}
            upd = [['label', text]]
            l_dic.update(upd)
            self.labl_dic[wid].update(l_dic)

    @pyqtSlot(int, str)
    def update_dir_dictionary(self, wid, dir: str):
        ''' Updates label dictionary with dirs to save images. '''
        if dir == '':
            return
        upd = [['dir', dir]]
        self.labl_dic[wid].update(upd)

    @pyqtSlot()
    def add_new_button(self):
        '''
        Func that adds new qwidget with buttons to settings,
        And label buttons to buttons widget.
        '''
        self.counter_new_button_id += 1
        self.sett_widget = Settings_widget(self.counter_new_button_id)
        self.butt_widget = Label_buttons(self.counter_new_button_id)
        self.sett_widget.text_emit.connect(self.create_dictionary)
        self.sett_widget.text_emit.connect(self.butt_widget.renaming_slot_button)
        self.sett_widget.dir_emit.connect(self.update_dir_dictionary)
        self.butt_widget.label_translate.connect(self.csvCreator)
        self.rename.connect(self.sett_widget.renaming_slot)
        self.docked_widget.layout().addWidget(self.sett_widget)
        self.sett_widget.delete.connect(self.delete_widget)

        self.docked_widget_2.layout().addWidget(self.butt_widget)

    @pyqtSlot(int)
    def delete_widget(self, wid: int):
        ''' Func that works with pyqtSignal that delets corresponded widget. '''
        self.counter_new_button_id -= 1

        widget = self.sender()
        self.docked_widget.layout().removeWidget(widget)

        # Finding children widgets of labels widget
        set_name_child = self.docked_widget_2.findChildren(QWidget, str(wid))
        set_name_without_list = set_name_child[0]

        widget.deleteLater()
        set_name_without_list.deleteLater()

    def open_files(self):
        fileNames, _ = QFileDialog.getOpenFileNames(self, 'QFileDialog.getOpenFileName()', '',
                                                  'Images (*.png *.jpeg *.jpg *.bmp *.gif)')
        self.names = fileNames

        if fileNames:  # Count files in directory
            self.path = os.path.dirname(fileNames[-1])
            for path in os.scandir(self.path):
                if path.is_file():
                    self.count += 1

        if fileNames:
            image = QImage(self.names[-1])
            if image.isNull():
                # QMessageBox.information(self, 'Image Viewer', 'Cannot load %s.' % fileName)
                return

            self.imageLabel.setPixmap(QPixmap.fromImage(image))
            self.scaleFactor = 0.0

            self.scroll_area.setVisible(True)
            self.image_label.adjustSize()

    def open_right(self):
        ''' Opens image in right spot. '''
        fileName, _ = QFileDialog.getOpenFileName(self, 'QFileDialog.getOpenFileName()', '',
                                                  'Images (*.png *.jpeg *.jpg *.bmp *.gif)')
        self.name = fileName

        if fileName:
            image = QImage(fileName)
            if image.isNull():
                QMessageBox.information(self, 'Image Viewer', 'Cannot load %s.' % fileName)
                return

            self.imageLabel_r.setPixmap(QPixmap.fromImage(image))
            self.scaleFactor = 0.0

            self.scroll_area_2.setVisible(True)
            self.image_label_r.adjustSize()

    def save_from_csv(self, destination: int):
        '''
        Saves img to desired dir.
        If not set up, saves into output_img_set folder.
        '''
        if 'dir' not in self.labl_dic.get(destination, {}):
            current_dir = self.output_img_set
        else:
            current_dir = self.labl_dic[destination]['dir']

        file_name = self.names.pop()
        naming = self.labl_dic[destination]['label']

        if file_name:
            path = os.path.join(current_dir, naming + f'{self.count_button_1}.png')
            self.path = path
            self.image_label.pixmap().save(path)
            self.count_button_1 += 1
        else:
            pass

        if len(self.names) == 0:
            QMessageBox.information(self, 'Image viewer', 'Images out!')
            return
        else:
            image = QImage(self.names[-1])

            if image.isNull():
                # QMessageBox.information(self, 'Image Viewer', 'Cannot load %s.' % fileName)
                return

            self.imageLabel.setPixmap(QPixmap.fromImage(image))
            self.scaleFactor = 0.0

            self.scroll_area.setVisible(True)
            self.image_label.adjustSize()

    def csvCreator(self, destination: int):
        '''
        Function to create table with image names (paths for now) and lables.
        And saving image.
        For now saves table to a current dir.
        '''
        ws = self.wb.active

        if destination in self.labl_dic:
            if self.count >= 2:
                self.save_from_csv(destination)
                self.row_count += 1
                cell = ws.cell(row=self.row_count, column=1)
                cell.value = str(self.path)
                cell = ws.cell(row=self.row_count, column=destination+1)
                cell.value = self.labl_dic[destination]['label']
                self.wb.save('img_data.xlsx')
            else:
                self.wb.save('img_data.xlsx')
        else:
            return


if __name__ == '__main__':

    app = QApplication(sys.argv)
    window = Image_Procession()
    window.show()
    sys.exit(app.exec())
